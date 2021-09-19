# coding=utf-8
import pandas as pd
from openpyxl import load_workbook
import json


if __name__ == '__main__':

    counter=1

    Reactions = load_workbook(filename="RemoodRawTest.xlsx")
    Data = load_workbook(filename="RemoodCleanTest.xlsx")
    
    ws = Reactions.active
    Dataws = Data.active
    
    ReadComments = ws['D']
    ReadShares=ws['E']
    ReadReactions=ws["F"]
    ReadLikes = ws['G']
    ReadLove = ws['H']
    ReadWow=ws['I']
    ReadLaugh = ws['J']
    ReadSad = ws['K']
    ReadAngry=ws['L']
    parent_id=ws["A"]
    

    
    Post_id = Dataws['C']

    
    Comments = {}
    shares={}
    totalRecations = {}
    likes = {}
    love = {}
    wow = {}
    haha = {}
    sad = {}
    angry = {}


    for row in range(1,len(parent_id)):
        Comments[str(parent_id[row].value)] = ReadComments[row].value
        shares[str(parent_id[row].value)]= ReadShares[row].value
        totalRecations[str(parent_id[row].value)] = ReadReactions[row].value
        likes[str(parent_id[row].value)] = ReadLikes[row].value
        love[str(parent_id[row].value)] = ReadLove[row].value
        wow[str(parent_id[row].value)] = ReadWow[row].value
        haha[str(parent_id[row].value)] = ReadLaugh[row].value
        sad[str(parent_id[row].value)]=ReadSad[row].value
        angry[str(parent_id[row].value)] = ReadAngry[row].value
        



    for key, data in Comments.items():
        while key != str(Post_id[counter].value):
           counter += 1
        Dataws.cell(column=15, row=counter + 1, value=data)
    counter=1

    for key, data in shares.items():
        while key != str(Post_id[counter].value):
           counter += 1
        Dataws.cell(column=16, row=counter + 1, value=data)
    counter=1

    for key, data in totalRecations.items():
        while key != str(Post_id[counter].value):
           counter += 1
        Dataws.cell(column=17, row=counter + 1, value=data)
    counter=1

    for key, data in likes.items():
        while key != str(Post_id[counter].value):
           counter += 1
        Dataws.cell(column=18, row=counter + 1, value=data)
    counter=1

    for key, data in love.items():
        while key != str(Post_id[counter].value):
           counter += 1
        Dataws.cell(column=19, row=counter + 1, value=data)
    counter=1

    for key, data in wow.items():
        while key != str(Post_id[counter].value):
           counter += 1
        Dataws.cell(column=20, row=counter + 1, value=data)
    counter=1

    for key, data in haha.items():
        while key != str(Post_id[counter].value):
           counter += 1
        Dataws.cell(column=21, row=counter + 1, value=data)
    counter = 0
    for key, data in sad.items():
        while key != str(Post_id[counter].value):
           counter += 1
        Dataws.cell(column=22, row=counter + 1, value=data)
    counter=1

    for key, data in angry.items():
        while key != str(Post_id[counter].value):
           counter += 1
        Dataws.cell(column=23, row=counter +1, value=data)
    counter=1



    Data.save("RemoodCleanDemo.xlsx")
